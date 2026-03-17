import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

DATABASE = 'database.db'
OUTPUT   = f'casamiento_datos_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'

# Paleta acorde a la web
COLOR_GOLD   = 'B8922A'
COLOR_BEIGE  = 'E8D9C0'
COLOR_DARK   = '1A1410'
COLOR_WHITE  = 'FFFFFF'
COLOR_LIGHT  = 'FAF5ED'

def estilo_header(cell, bg=COLOR_GOLD):
    cell.font      = Font(bold=True, color=COLOR_WHITE, name='Arial', size=11)
    cell.fill      = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border    = borde_fino()

def estilo_titulo(cell):
    cell.font      = Font(bold=True, color=COLOR_DARK, name='Arial', size=13)
    cell.fill      = PatternFill('solid', start_color=COLOR_BEIGE)
    cell.alignment = Alignment(horizontal='center', vertical='center')

def estilo_fila(cell, impar=True):
    cell.fill      = PatternFill('solid', start_color=COLOR_LIGHT if impar else COLOR_WHITE)
    cell.font      = Font(name='Arial', size=10)
    cell.alignment = Alignment(vertical='center', wrap_text=True)
    cell.border    = borde_fino()

def borde_fino():
    lado = Side(style='thin', color='C9B99A')
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def autofit(ws, columnas):
    for col, ancho in columnas.items():
        ws.column_dimensions[col].width = ancho

def escribir_tabla(ws, titulo, headers, filas, anchos):
    # Título de la hoja
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    titulo_cell = ws['A1']
    titulo_cell.value = titulo
    estilo_titulo(titulo_cell)
    ws.row_dimensions[1].height = 30

    # Headers
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        estilo_header(cell)
    ws.row_dimensions[2].height = 22

    # Datos
    if filas:
        for i, fila in enumerate(filas):
            for col, valor in enumerate(fila, 1):
                cell = ws.cell(row=i+3, column=col, value=valor)
                estilo_fila(cell, impar=(i % 2 == 0))
            ws.row_dimensions[i+3].height = 20
    else:
        ws.merge_cells(f'A3:{get_column_letter(len(headers))}3')
        cell = ws['A3']
        cell.value     = 'Sin registros aún.'
        cell.font      = Font(italic=True, color='7A6A5A', name='Arial')
        cell.alignment = Alignment(horizontal='center')

    # Fila de total (solo si hay datos)
    if filas:
        total_row = len(filas) + 3
        ws.merge_cells(f'A{total_row}:{get_column_letter(len(headers)-1)}{total_row}')
        cell_label = ws[f'A{total_row}']
        cell_label.value     = 'Total de registros'
        cell_label.font      = Font(bold=True, name='Arial', size=10)
        cell_label.alignment = Alignment(horizontal='right')
        cell_label.fill      = PatternFill('solid', start_color=COLOR_BEIGE)

        cell_total = ws.cell(row=total_row, column=len(headers))
        cell_total.value = f'=COUNTA({get_column_letter(len(headers))}3:{get_column_letter(len(headers))}{total_row-1})'
        cell_total.font  = Font(bold=True, name='Arial', size=10)
        cell_total.fill  = PatternFill('solid', start_color=COLOR_BEIGE)
        cell_total.alignment = Alignment(horizontal='center')

    autofit(ws, anchos)

def exportar():
    conn   = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Eliminar hoja vacía por defecto

    # ── Hoja 1: RSVP ──────────────────────────────────────
    ws_rsvp = wb.create_sheet('Asistencia')
    cursor.execute('SELECT nombre, dni, restriccion, telefono, fecha_registro FROM rsvp ORDER BY fecha_registro DESC')
    filas_rsvp = cursor.fetchall()
    escribir_tabla(
        ws   = ws_rsvp,
        titulo  = '✉ Confirmaciones de Asistencia — Silvina & Maximiliano',
        headers = ['Nombre', 'DNI', 'Restricción alimentaria', 'Teléfono', 'Fecha de registro'],
        filas   = filas_rsvp,
        anchos  = {'A': 28, 'B': 16, 'C': 14, 'D': 18, 'E': 22}
    )

    # ── Hoja 2: Música ────────────────────────────────────
    ws_musica = wb.create_sheet('Sugerencias Música')
    cursor.execute('SELECT nombre, cancion, artista, fecha_registro FROM musica ORDER BY fecha_registro DESC')
    filas_musica = cursor.fetchall()
    escribir_tabla(
        ws      = ws_musica,
        titulo  = '🎵 Sugerencias de Canciones — Silvina & Maximiliano',
        headers = ['Nombre', 'Canción', 'Artista', 'Fecha de registro'],
        filas   = filas_musica,
        anchos  = {'A': 24, 'B': 30, 'C': 24, 'D': 22}
    )

    # ── Hoja 3: Cápsula de deseos ─────────────────────────
    ws_carta = wb.create_sheet('Cápsula de Deseos')
    cursor.execute('SELECT nombre, sugerencia, fecha_registro FROM carta ORDER BY fecha_registro DESC')
    filas_carta = cursor.fetchall()
    escribir_tabla(
        ws      = ws_carta,
        titulo  = '💌 Cápsula de Deseos — Silvina & Maximiliano',
        headers = ['Nombre', 'Mensaje', 'Fecha de registro'],
        filas   = filas_carta,
        anchos  = {'A': 24, 'B': 60, 'C': 22}
    )

    conn.close()
    wb.save(OUTPUT)
    print(f'✅ Exportado correctamente: {OUTPUT}')
    print(f'   → {len(filas_rsvp)} confirmaciones de asistencia')
    print(f'   → {len(filas_musica)} sugerencias de canciones')
    print(f'   → {len(filas_carta)} mensajes en la cápsula')

if __name__ == '__main__':
    exportar()